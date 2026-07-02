from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from database import get_conn

COLUMNS = [
    ("fecha", "Fecha"),
    ("idioma", "Idioma"),
    ("unidad_long", "Unidad Longitud"),
    ("unidad_peso", "Unidad Peso"),
    ("proveedor_codigo", "Código Proveedor"),
    ("proveedor_nombre", "Nombre Proveedor"),
    ("proveedor_direccion", "Dirección Proveedor"),
    ("proveedor_contacto", "Contacto Proveedor"),
    ("proveedor_email", "Email Proveedor"),
    ("proveedor_telefono", "Teléfono Proveedor"),
    ("numero_pieza", "N° Pieza"),
    ("descripcion", "Descripción Pieza"),
    ("largo", "Largo Pieza"),
    ("ancho", "Ancho Pieza"),
    ("alto", "Alto Pieza"),
    ("peso", "Peso Pieza"),
    ("moq", "MOQ"),
    ("proyecto", "Proyecto"),
    ("emb_identico", "Embalaje Idéntico (1/0)"),
    ("p1_tipo", "P1 Tipo"),
    ("p1_retornable", "P1 Retornable"),
    ("p1_descripcion", "P1 Descripción"),
    ("p1_largo", "P1 Largo"),
    ("p1_ancho", "P1 Ancho"),
    ("p1_alto", "P1 Alto"),
    ("p1_peso_emb", "P1 Peso Embalaje"),
    ("p1_capacidad", "P1 Capacidad"),
    ("p1_peso_bruto", "P1 Peso Bruto"),
    ("p2_tipo", "P2 Tipo"),
    ("p2_retornable", "P2 Retornable"),
    ("p2_descripcion", "P2 Descripción"),
    ("p2_largo", "P2 Largo"),
    ("p2_ancho", "P2 Ancho"),
    ("p2_alto", "P2 Alto"),
    ("p2_peso_emb", "P2 Peso Embalaje"),
    ("p2_capacidad", "P2 Capacidad"),
    ("p2_peso_bruto", "P2 Peso Bruto"),
    ("accesorios", "Accesorios"),
    ("ap_elaborado_nombre", "Elaborado Por"),
    ("ap_elaborado_fecha", "Fecha Elaborado"),
    ("ap_revisado_nombre", "Revisado Por"),
    ("ap_revisado_fecha", "Fecha Revisado"),
    ("ap_aprobado_nombre", "Aprobado Por"),
    ("ap_aprobado_fecha", "Fecha Aprobado"),
    ("estado", "Estado"),
]

def export_riai_excel():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM riai ORDER BY id DESC").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "RIAIs"

    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    ws.append(["ID"] + [label for _, label in COLUMNS])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        row_data = [r["id"]] + [r[key] if r[key] is not None else "" for key, _ in COLUMNS]
        ws.append(row_data)

    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_proveedores_excel():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM proveedores").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    headers = ["Código", "Nombre", "Dirección", "Contacto", "Email", "Teléfono"]
    ws.append(headers)
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    for r in rows:
        ws.append([r["codigo"], r["nombre"], r["direccion"], r["contacto"], r["email"], r["telefono"]])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_riai_excel(file_stream):
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    col_map = {}
    for key, label in COLUMNS:
        if label in headers:
            col_map[key] = headers.index(label)

    conn = get_conn()
    inserted = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue
        try:
            data = {}
            for key, _ in COLUMNS:
                idx = col_map.get(key)
                val = row[idx] if idx is not None and idx < len(row) else None
                data[key] = val if val is not None else ""
            placeholders = ", ".join(f":{k}" for k, _ in COLUMNS)
            cols = ", ".join(k for k, _ in COLUMNS)
            conn.execute(f"INSERT INTO riai ({cols}) VALUES ({placeholders})", data)
            inserted += 1
        except Exception as e:
            errors.append(f"Fila {row_idx}: {str(e)}")

    conn.commit()
    conn.close()
    return inserted, errors


def import_proveedores_excel(file_stream):
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active

    conn = get_conn()
    inserted = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue
        try:
            codigo, nombre, direccion, contacto, email, telefono = (list(row) + [None]*6)[:6]
            if not codigo or not nombre:
                errors.append(f"Fila {row_idx}: falta código o nombre")
                continue
            conn.execute("""
                INSERT INTO proveedores (codigo, nombre, direccion, contacto, email, telefono)
                VALUES (?,?,?,?,?,?)
                ON CONFLICT(codigo) DO UPDATE SET
                    nombre=excluded.nombre, direccion=excluded.direccion,
                    contacto=excluded.contacto, email=excluded.email, telefono=excluded.telefono
            """, (codigo, nombre, direccion, contacto, email, telefono))
            inserted += 1
        except Exception as e:
            errors.append(f"Fila {row_idx}: {str(e)}")

    conn.commit()
    conn.close()
    return inserted, errors
